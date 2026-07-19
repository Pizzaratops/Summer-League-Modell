#!/usr/bin/env python3
"""
update_data.py
──────────────
Lädt die komplette Google-Sheet-Arbeitsmappe (als .xlsx-Export) herunter und
baut daraus data/data.json komplett neu auf, damit die Webseite (index.html)
automatisch aktuelle Zahlen anzeigt.

Wird 3x täglich von .github/workflows/update-data.yml ausgeführt.

Struktur des Sheets (ein Tab pro Team + FA26 + Finances + CtrlPanel):
  - Team-Tabs (ATL, BOS, ... WA):
      Zeile 1: Team-Vollname
      Zeile 2: Header (GM: ..., CAP HIT, 2026-27, 2027-28, 2028-29, 2029-30,
               2030-31, 2031-32, QO, PG, SG, SF, PF, C, TRADABLE ALONE,
               TRADABLE AGGREGATED, SIGNED USING, LB, EB, NB, NOTES)
      Zeile 3+: Spieler (aktiver Kader, Two-way Players, Waived Players),
                Spalte A der 3. Zeile enthält zusätzlich die "direction"
                (z.B. "Rebuild", "Contender", ...)
      Danach eine Zeile "$ | Total salaries | <Summe>"
  - Finances-Tab: 'Owner cap' und 'Left under owner cap' pro Team
  - CtrlPanel-Tab: Salary Cap / Tax / Aprons / Average Salary (Label in
    Spalte D, Wert in Spalte E)
  - FA26-Tab: alle relevanten Spieler mit 25/26 FPG-Wert; "echte" Free
    Agents sind alle, die in KEINEM Team-Tab als Spieler auftauchen.

Team-Farben, Vollnamen und Coach-Namen stehen nicht im Sheet und werden
daher aus TEAM_META (unten) übernommen bzw. aus der bestehenden
data/data.json beibehalten, falls dort schon etwas gepflegt wurde.
"""

import io
import json
import os
import re
import sys
import urllib.request
from datetime import datetime, timezone

import openpyxl

# ── KONFIGURATION ────────────────────────────────────────────────────────
SHEET_ID = "1REnOtl5b7IQSbxBaeGPPs6KnPwhHmwY8vltQcVh6jgQ"

DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "data.json")

TEAM_ABBRS = [
    "ATL", "BOS", "BRK", "CHA", "CHI", "CLE", "DAL", "DEN", "DET", "GS",
    "HOU", "IND", "LAC", "LAL", "MEM", "MIA", "MIL", "MIN", "NOP", "NY",
    "OKC", "ORL", "PHI", "PHX", "POR", "SAC", "SAS", "TOR", "UT", "WA",
]

# Spalten (0-indiziert) mit Positions-Flags in den Team-Tabs.
POS_COLS = [10, 11, 12, 13, 14]
POS_LABELS = ["PG", "SG", "SF", "PF", "C"]

# Statische Team-Infos, die NICHT im Sheet stehen (Farben, Vollname, Coach).
# Werden als Fallback benutzt bzw. mit bestehender data.json gemerged.
TEAM_META = {
    "ATL": {"name": "Atlanta Hawks", "c1": "#E03A3E", "c2": "#C1D32F", "coach": ""},
    "BOS": {"name": "Boston Celtics", "c1": "#007A33", "c2": "#BA9653", "coach": ""},
    "BRK": {"name": "Brooklyn Nets", "c1": "#000000", "c2": "#AAAAAA", "coach": ""},
    "CHA": {"name": "Charlotte Hornets", "c1": "#1D1160", "c2": "#00788C", "coach": "Stan Van Gundy"},
    "CHI": {"name": "Chicago Bulls", "c1": "#CE1141", "c2": "#000000", "coach": ""},
    "CLE": {"name": "Cleveland Cavaliers", "c1": "#860038", "c2": "#FDBB30", "coach": ""},
    "DAL": {"name": "Dallas Mavericks", "c1": "#00538C", "c2": "#B8C4CA", "coach": "Rick Carlisle"},
    "DEN": {"name": "Denver Nuggets", "c1": "#0E2240", "c2": "#FEC524", "coach": "Andre Miller"},
    "DET": {"name": "Detroit Pistons", "c1": "#C8102E", "c2": "#006BB6", "coach": ""},
    "GS": {"name": "Golden State Warriors", "c1": "#1D428A", "c2": "#FFC72C", "coach": ""},
    "HOU": {"name": "Houston Rockets", "c1": "#CE1141", "c2": "#C4CED4", "coach": ""},
    "IND": {"name": "Indiana Pacers", "c1": "#002D62", "c2": "#FDBB30", "coach": "Rick Carlisle"},
    "LAC": {"name": "LA Clippers", "c1": "#C8102E", "c2": "#1D428A", "coach": ""},
    "LAL": {"name": "Los Angeles Lakers", "c1": "#552583", "c2": "#FDB927", "coach": ""},
    "MEM": {"name": "Memphis Grizzlies", "c1": "#5D76A9", "c2": "#12173F", "coach": ""},
    "MIA": {"name": "Miami Heat", "c1": "#98002E", "c2": "#F9A01B", "coach": ""},
    "MIL": {"name": "Milwaukee Bucks", "c1": "#00471B", "c2": "#EEE1C6", "coach": ""},
    "MIN": {"name": "Minnesota Timberwolves", "c1": "#0C2340", "c2": "#236192", "coach": ""},
    "NOP": {"name": "New Orleans Pelicans", "c1": "#0C2340", "c2": "#C8102E", "coach": ""},
    "NY": {"name": "New York Knicks", "c1": "#006BB6", "c2": "#F58426", "coach": "Quin Snyder"},
    "OKC": {"name": "OKC Thunder", "c1": "#007AC1", "c2": "#EF3B24", "coach": ""},
    "ORL": {"name": "Orlando Magic", "c1": "#0077C0", "c2": "#C4CED4", "coach": ""},
    "PHI": {"name": "Philadelphia 76ers", "c1": "#006BB6", "c2": "#ED174C", "coach": "Ime Udoka"},
    "PHX": {"name": "Phoenix Suns", "c1": "#1D1160", "c2": "#E56020", "coach": "Jeff Hornacek"},
    "POR": {"name": "Portland Trail Blazers", "c1": "#E03A3E", "c2": "#000000", "coach": ""},
    "SAC": {"name": "Sacramento Kings", "c1": "#5A2D81", "c2": "#63727A", "coach": ""},
    "SAS": {"name": "San Antonio Spurs", "c1": "#8A8D8F", "c2": "#000000", "coach": ""},
    "TOR": {"name": "Toronto Raptors", "c1": "#CE1141", "c2": "#000000", "coach": "Jeff Van Gundy"},
    "UT": {"name": "Utah Jazz", "c1": "#002B5C", "c2": "#F9A01B", "coach": ""},
    "WA": {"name": "Washington Wizards", "c1": "#002B5C", "c2": "#E31837", "coach": ""},
}

# Team-Vollname (wie er im Finances-Tab steht) → Abkürzung.
NAME_TO_ABBR = {
    "Atlanta Hawks": "ATL", "Boston Celtics": "BOS", "Brooklyn Nets": "BRK",
    "Charlotte Hornets": "CHA", "Chicago Bulls": "CHI", "Cleveland Cavaliers": "CLE",
    "Dallas Mavericks": "DAL", "Denver Nuggets": "DEN", "Detroit Pistons": "DET",
    "Golden State Warriors": "GS", "Houston Rockets": "HOU", "Indiana Pacers": "IND",
    "LA Clippers": "LAC", "Los Angeles Clippers": "LAC", "Los Angeles Lakers": "LAL",
    "Memphis Grizzlies": "MEM", "Miami Heat": "MIA", "Milwaukee Bucks": "MIL",
    "Minnesota Timberwolves": "MIN", "New Orleans Pelicans": "NOP",
    "New York Knicks": "NY", "OKC Thunder": "OKC", "Oklahoma City Thunder": "OKC",
    "Orlando Magic": "ORL", "Philadelphia 76ers": "PHI", "Phoenix Suns": "PHX",
    "Portland Trail Blazers": "POR", "Sacramento Kings": "SAC",
    "San Antonio Spurs": "SAS", "Toronto Raptors": "TOR", "Utah Jazz": "UT",
    "Washington Wizards": "WA",
}

# Default my_teams, falls in der bestehenden data.json noch nichts steht.
DEFAULT_MY_TEAMS = ["PHI", "UT"]


def workbook_export_url(sheet_id: str) -> str:
    # Lädt die GESAMTE Arbeitsmappe (alle Tabs) als .xlsx - keine gid nötig.
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def fetch_workbook(url: str) -> openpyxl.Workbook:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        raw = resp.read()
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)


def num(v):
    """Wandelt einen Zellwert robust in int um (oder None)."""
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(str(v).replace("$", "").replace(",", "").strip()))
        except (TypeError, ValueError):
            return None


def parse_gm(raw) -> str:
    if not raw:
        return ""
    text = re.sub(r"(?i)^\s*GM:\s*", "", str(raw))
    return text.split(",")[0].strip()


def parse_team_sheet(ws):
    """Liest einen Team-Tab und gibt (direction, gm, players, total_salary) zurück."""
    direction_raw = ws.cell(row=3, column=1).value
    direction = direction_raw.strip() if isinstance(direction_raw, str) else ""
    gm = parse_gm(ws.cell(row=2, column=1).value)

    players = []
    total_salary = None
    for r in range(3, ws.max_row + 1):
        col1 = ws.cell(row=r, column=2).value
        if isinstance(col1, str) and col1.strip() == "Total salaries":
            # Spalte C = "CAP HIT", Spalte D = "2026-27" (aktuelle Saison).
            # Für Cap-/Apron-Vergleiche muss die aktuelle Saison genommen werden,
            # nicht der Cap Hit (der z.B. Signing-Bonus-Verteilung o.ä. enthalten kann).
            total_salary = num(ws.cell(row=r, column=4).value)
            break
        if not isinstance(col1, str) or not col1.strip():
            continue

        name = col1.strip()
        row = [ws.cell(row=r, column=c).value for c in range(1, 23)]

        pos_flags = [POS_LABELS[i] for i, c in enumerate(POS_COLS) if row[c] not in (None, "")]
        pos = "/".join(pos_flags) if pos_flags else "?"

        bird = None
        if row[18] not in (None, ""):
            bird = "Full"
        elif row[19] not in (None, ""):
            bird = "Early"
        elif row[20] not in (None, ""):
            bird = "Non"

        signed_via = row[17].strip() if isinstance(row[17], str) and row[17].strip() else None

        players.append({
            "name": name,
            "pos": pos,
            "sal_26": num(row[3]),
            "sal_27": num(row[4]),
            "sal_28": num(row[5]),
            "sal_29": num(row[6]),
            "sal_30": num(row[7]),
            "bird": bird,
            "signed_via": signed_via,
            "tradable_alone": bool(row[15]),
            "tradable_agg": bool(row[16]),
        })

    if total_salary is None:
        total_salary = sum(p["sal_26"] or 0 for p in players)

    return direction, gm, players, total_salary


def parse_finances(wb) -> dict:
    ws = wb["Finances"]
    header = [c.value for c in ws[1]]
    try:
        idx_owner_cap = header.index("Owner cap")
        idx_left = header.index("Left under owner cap")
    except ValueError:
        print("  ⚠️ 'Finances'-Tab hat unerwartete Spalten, überspringe Owner Caps.", file=sys.stderr)
        return {}

    owner_caps = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        team_name = row[0]
        if not team_name:
            continue
        abbr = NAME_TO_ABBR.get(str(team_name).strip())
        if not abbr:
            continue
        owner_caps[abbr] = {
            "owner_cap": num(row[idx_owner_cap]),
            "left": num(row[idx_left]),
        }
    return owner_caps


def parse_ctrlpanel(wb) -> dict:
    ws = wb["CtrlPanel"]
    label_map = {
        "salary cap": "cap",
        "luxury tax threshold": "tax",
        "first apron": "apron1",
        "second apron": "apron2",
        "average salary": "avgSal",
    }
    cba = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 5:
            continue
        label = row[3]
        if isinstance(label, str) and label.strip().lower() in label_map:
            cba[label_map[label.strip().lower()]] = num(row[4])
    return cba


def parse_fa26(wb, rostered: dict) -> list:
    """Liest den FA26-Tab komplett aus.

    `rostered` ist ein dict {Spielername: Team-Abkürzung} für alle Spieler,
    die auf einem Team-Roster stehen. Statt diese Spieler auszufiltern,
    markieren wir sie mit `signed_team`, damit man auf der Webseite sieht:
    "steht zwar im FA-Pool des Sheets, ist aber schon vergeben".
    """
    ws = wb["FA26"]
    fas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not isinstance(name, str) or not name.strip():
            continue
        name = name.strip()
        fpg_raw = row[2] if len(row) > 2 else None
        try:
            fpg = round(float(fpg_raw), 2) if fpg_raw is not None else None
        except (TypeError, ValueError):
            fpg = None
        fas.append({
            "name": name,
            "fpg": fpg,
            "signed_team": rostered.get(name),
        })
    fas.sort(key=lambda p: (p["fpg"] is None, -(p["fpg"] or 0)))
    return fas


def build_data(wb, existing: dict) -> dict:
    teams = {}
    directions = {}
    rostered = {}  # Spielername -> Team-Abkürzung

    for abbr in TEAM_ABBRS:
        if abbr not in wb.sheetnames:
            print(f"  ⚠️ Tab '{abbr}' nicht im Sheet gefunden, überspringe.", file=sys.stderr)
            continue
        ws = wb[abbr]
        direction, gm, players, total_salary = parse_team_sheet(ws)
        directions[abbr] = direction
        meta = TEAM_META.get(abbr, {})
        existing_team = existing.get("teams", {}).get(abbr, {})
        teams[abbr] = {
            "abbr": abbr,
            "name": meta.get("name", existing_team.get("name", abbr)),
            "c1": meta.get("c1", existing_team.get("c1", "#888888")),
            "c2": meta.get("c2", existing_team.get("c2", "#CCCCCC")),
            "coach": existing_team.get("coach", meta.get("coach", "")),
            "gm": gm,
            "players": players,
            "total_salary": total_salary,
        }
        for p in players:
            rostered[p["name"]] = abbr

    owner_caps = parse_finances(wb)
    cba = parse_ctrlpanel(wb)
    fas = parse_fa26(wb, rostered)

    my_teams = existing.get("my_teams") or DEFAULT_MY_TEAMS

    return {
        "directions": directions,
        "owner_caps": owner_caps,
        "teams": teams,
        "fas": fas,
        "my_teams": my_teams,
        "cba": cba or existing.get("cba", {}),
        "meta": {
            "last_updated": datetime.now(timezone.utc).isoformat(),
            "source": "Google Sheets (auto-fetch)",
            "teams_count": len(teams),
            "players_count": sum(len(t["players"]) for t in teams.values()),
            "fas_count": len(fas),
            "fas_available_count": sum(1 for f in fas if not f["signed_team"]),
        },
    }


def main() -> int:
    existing = {}
    if os.path.exists(DATA_PATH):
        with open(DATA_PATH, encoding="utf-8") as f:
            existing = json.load(f)

    url = workbook_export_url(SHEET_ID)
    print(f"→ Lade Arbeitsmappe von {url}")
    try:
        wb = fetch_workbook(url)
    except Exception as e:
        print(f"❌ Fehler beim Laden der Arbeitsmappe: {e}", file=sys.stderr)
        return 1

    print(f"  {len(wb.sheetnames)} Tabs gefunden.")
    data = build_data(wb, existing)

    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    n_teams = len(data["teams"])
    n_players = sum(len(t["players"]) for t in data["teams"].values())
    print(f"✅ data/data.json aktualisiert: {n_teams} Teams, {n_players} Spieler, {len(data['fas'])} Free Agents.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
